import openpyxl
from openpyxl.styles import PatternFill


def getrowcount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def getcolumncount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def readdata(file, sheetname, rownum, colnumno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum, colnumno).value


def writedata(file, sheetname, rownum, colnumno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum, colnumno).value = data
    workbook.save(file)


def fillgreencolor(file, sheetname, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenfill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum, columnno).fill = greenfill
    workbook.save(file)


def fillredcolor(file, sheetname, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redfill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rownum, columnno).fill = redfill
    workbook.save(file)
